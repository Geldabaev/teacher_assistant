import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd


__all__ = ['main_exel', 'calculate_salary']


def create_excel():
    """Создаем документ xlsx если нет"""
    cur_date = datetime.now().strftime("%d_%m_%Y")

    try:
        # если файл есть дописываем
        wb = openpyxl.load_workbook("excel/write_only.xlsx")
    except:
        # если нет создаем
        wb = openpyxl.Workbook()
        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            wb.remove(sheet)

            # создаем новый лист
            ws = wb.create_sheet(f'{cur_date}', 0)
            wb.save('excel/write_only.xlsx')


    # active берет первый лист
    sheet = wb.active

    if sheet['A1'].value is None:  # Пишем заголовок если нет
        sheet['A1'].value = "Группы"
        wb.save('excel/write_only.xlsx')

    return cur_date


def search_group(name_group):
    'поиск группы, если есть запись'
    try:
        wb = openpyxl.open("excel/write_only.xlsx")
    except:
        print('Записи ещё нет!')

    sheet = wb.active
    # ищем группу
    # должно соответствовать переданное в функцию имя
    # ходим по всем заявкам ищя название
    for row in range(2, sheet.max_row + 1):
        if sheet[f'A{row}'].value == name_group:
            return (row, sheet, wb)  # Если групп найдена возвращаем его местоположение

    return (False, sheet, wb)  # Если группы нет, записываем в пустое поле и возвращем его местоположение


def write_group(sheet, name_group, wb):
    """сохраняем данные"""
    # запишем данные
    # с пустого ряда
    row = 2  # начинаем поиск пустого ряда с ряда номер 1
    while True:
        # Вставляем название группы в пустое поле
        if sheet[f'A{row}'].value is None:
            sheet[f'A{row}'].value = name_group
            wb.save('excel/write_only.xlsx')
            return row  # возвращаем местоположение группы
        row += 1


def werite_count_students(sheet, row, count_, wb, cur_data):
    for i in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']:
        if sheet[f'{i}{1}'].value == cur_data:
            sheet[f'{i}{row}'].value = count_
            break
        elif sheet[f'{i}{1}'].value is None:  # ищем пустуя ячейку для записи кол. студентов

            """Нужно проверить есть ли такая дата:
                    если есть пишем в этот столбек в пустую ячейку в поле где группа
                    если нет создаем новую дату в пустой ячейке там где заголовки"""
            sheet[f'{i}{1}'].value = cur_data
            sheet[f'{i}{row}'].value = count_
            break

    wb.save('excel/write_only.xlsx')


def main_exel(name_group, count_):
    cur_date = create_excel()
    res_row_sheet = search_group(name_group)
    if all(res_row_sheet):  # такая группа найдена
        werite_count_students(res_row_sheet[1], res_row_sheet[0], count_, res_row_sheet[2], cur_date)
    else:
        res_row = write_group(res_row_sheet[1], name_group, res_row_sheet[2])
        werite_count_students(res_row_sheet[1], res_row, count_, res_row_sheet[2], cur_date)


def calculate_salary():
    df = pd.read_excel("../zp/excel/write_only.xlsx")
    df.fillna(0, inplace=True)
    zp = 0
    for i in df.dropna().values:
        if not i[0].startswith("Django"):
            for a in i[1:]:
                if a > 0:
                    if a <= 5:
                        zp += 500
                    else:
                        zp += a * 100
        else:
            for a in i[1:]:
                if a > 0:
                    if a <= 5:
                        zp += 1000
                    else:
                        zp += a * 200
    return zp



# main_exel("Габатаевы группа", 1)
